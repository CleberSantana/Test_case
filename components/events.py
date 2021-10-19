import os
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

l = []


def sheetlist(file_path):
    try:
        sheetlist1 = {}
        wb = load_workbook(filename = os.path.basename(file_path), data_only = True)
        for i in wb.sheetnames:
            ws = wb[str(i)] 
            if ws.sheet_properties.tabColor.value == 'FFA5A5A5':
                sheetlist1[i] = wb[str(i)].cell(column = 9, row = 3).value
        wb.close()
        return sheetlist1   
    except:
        pass

   
def def_path(): # function for choose a file 
    root = tk.Tk() 
    root.withdraw()
    file_path = filedialog.askopenfilename(initialdir=os.getcwd(), title="Select file", filetypes=[("TestCase Files", "*.xlsx")])
    if file_path.endswith(".xlsx"):
        return os.path.abspath(file_path)
    else:
        return 'Select a file'
   
    
d = {}
def create_dict(title:int, dict1:dict): # function to create a dict 
    d[title] = dict1 # assing the spec and status to the dictionary
    return d # returning the dictionary


#creating list from cvs file
def filtering(file_path:str, currenttest:str):
    title = []
    d1 = {}
    d2 = {}
    c = 0
    dv = 0
    wb = load_workbook(filename = os.path.basename(file_path), data_only = True)
    ws = wb[str(currenttest)]
    totalrow = ws.max_row

    for row in ws.iter_rows(min_row = 5, max_row = totalrow+1, min_col = 3, max_col = 3):
        for cell in row:
            if ("Test" in str(cell.value) or cell.row == totalrow):
                if ws.cell(column = 5, row = cell.row).value is not None:
                    title.append(str(ws.cell(column = 5, row = cell.row).value))
                if c != 0:
                    d1 = d2.copy()
                    d = create_dict(title[dv], d1)
                    dv += 1
                    d2.clear()
            else:
                if ws.cell(column = 5, row = cell.row).value is None:
                    pass
                else:
                    d2[c] = {'step' : ws.cell(column = 5, row = cell.row).value.strip(),
                            'Expected' : ws.cell(column = 6, row = cell.row).value,
                            'Actual': ws.cell(column = 7, row = cell.row).value,
                            'sts': ws.cell(column = 8, row = cell.row).value,
                            'cellrow' : cell.row}
                    c += 1
    
    wb.close()
                    
    return d, title


def saving_data(file_path:str, procedure:str, sheet:str):
    
    wb = load_workbook(filename = os.path.basename(file_path), data_only = True)
    ws = wb[str(sheet)]
    
    for i in d.keys():
        for j in d[i].keys():
            ws.cell(column=7, row=int(d[i][j]['cellrow']), value=str(d[i][j]['Actual'].strip()))

    wb.save(filename = os.path.basename(file_path))
    wb.close()
        